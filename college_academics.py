#!/usr/bin/env python3

import os, sys
import re
import pickle
from config import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter


#FUNCTIONS  
def parse_college_name(college):
    names = { "IH": ("international", "internationalhouse", "international_house"),
                "FTH" : ("fth", "fig", "fig_tree", "fig_tree_hall"),
                "HALL" : ("unsw", "unswhall", "unsw_hall", "hall"),
                "BASS": ("basser", "bass"),
                "BAXT": ("baxter", "bax"),
                "GOLD": ("goldstein", "gold"),
                "COLH": ("colombo", "colombo_house", "colombohouse")
            }
    for key in names.keys():
        if college.lower() in names[key]:
            return key
    return college
  
def pdf_to_txt(directory):
    # preprocessing: convert pdfs in dir to txt files
        for file in os.listdir(directory):
            if file.endswith(".pdf"):
                os.system(f"pdftotext -layout {directory}/\'{file}\'")    

def parse_lines(fp, college, all_terms):
    # data dict
    students = {}
    
    # patterns for regex
    pattern = {
            "name_zid" : '\W?([\w\-\.\' ]+) \((\d{7})\)',
            "enrolment_history" : '(?i)^enrolment history details$',
            "type_program" : '(?i)^([\w\-]+)\s{5,}(\d{4}\s+\w.*)$',
            "term" : '(?i)^\s*((?:term|semester|summer).+?\d{4}\s*$)',
            # "course" : '(?i)^\s*(\w{4}) ?(\d{4})\s*((?:[\w\d\:\.,-\(\)]+ ?)+?)(?:\s{3,})?(\d{2,3})?\s+?([\w ]+)?'
            "course" : '(?i)^\s*(\w{4}) ?(\d{4})\s*',
            "course_tabs" : '^#?(\w{4}) ?(\d{4})#?((?:[\w\d\:\.,\-\(\)]+ ?)+)#?(\d{2,3})?#?([\w ]+)?'
            
        }
        
    #initialise object variable
    student = None    
    
   # parse loop
    line = fp.readline()
    while line:
        # Skip empty line
        if not line:
            line = fp.readline()
            continue
        
        # match name line
        match = re.search(pattern['name_zid'], line)
        if match:
            name = match.group(1)
            zid = match.group(2)
            
            # skip if start of additional page for existing entry
            if student and zid == student.zid:
                line = fp.readline()
                continue
            elif student:
                students[student.zid] = student
            
            #new student object
            student = Student(name.strip(), zid.strip(), college.strip())

        # continue from enrolment history details
        match = re.search(pattern['type_program'], line)
        if match:
            enrol_type = match.group(1)
            program = match.group(2)
            student.enrol_type = enrol_type.strip().title()
            student.program = program.strip()
        
        # process terms
        match = re.search(pattern['term'], line)
        if match:
            term = match.group().strip()
            term = term[-4:] + " " + term[:-5]
            all_terms.add(term)
            if term not in student.terms:
                student.terms[term] = []
            
        # process courses
        match = re.search(pattern['course'], line)
        if match:
            #
            match = re.sub('\s{3,}', '#', match.string)
            match = re.search(pattern['course_tabs'], match)
            #
            code_name = match.group(1)
            code_num = match.group(2)
            name = match.group(3)
            grade = match.group(4)
            grade_name = match.group(5)
            
            grade = '' if not grade else grade
            grade_name = '' if not grade_name else grade_name
            
            course = Course(*map(str.strip, [code_name, code_num, name, grade, grade_name]))
            student.addCourse(term, course)
        
        line = fp.readline()
    
    if student:
        students[student.zid] = student
        
    return students


                
# User to select term            
def pick_term(all_terms):
    # get term for excel processing
    all_terms = sorted(all_terms)
    print()
    for i, k in enumerate(all_terms, 1):
        print(f"{i}. {k} ({convert_term_name(k)})")
        
    terms = []
    while True:
        term_start = int(input("\nPlease select the academic term you wish to start processing from (e.g. 22):\n"))
        if term_start in range(1, len(all_terms)+2): 
            term_start -= 1
            break
        
    while True:
        term_end = input("\nPlease select the academic term you wish to end processing on (e.g. 24); hit enter if only processing one term:\n")
        if term_end == "":
            term_end = term_start
            break
        elif int(term_end) in range(1, len(all_terms)+2): 
            term_end = int(term_end)
            term_end -= 1
            break
    
    for t in range(term_start, term_end + 1):
        terms.append(all_terms[t])
        
    print(f"\nYou have selected: {', '.join(terms)}\n")
    return terms                
                
# Print a student dict to excel                
def export_data(ws, students_dict, term, export_all=False):
    # Initialise sheet headers
    col_headers = ["First Names", "Last Name", "zID", "College", "Type", "Program", "Term", "Code", "Course", "Mark", "Grade", "WAM"]
    for row in ws.iter_rows(min_row=1, max_col=len(col_headers), max_row=1):
        for i, cell in enumerate(row):
            cell.value = col_headers[i]
            cell.font = Font(bold = True)
           
    r = 2
    # print each student's data
    for zid, student in students_dict.items():
        # start = r
        min_rows = 3
        for c, (prop, val) in enumerate(student.__dict__.items(), 1):
            # print courses
            if prop == "terms": 
                if term in val.keys():
                    # print Term
                    ws.cell(row=r, column=c, value=term.title())
                    # print WAM
                    wam = student.wams[term]
                    if wam:
                        wam = float(student.wams[term])
                    ws.cell(row=r, column=c+5, value=wam)
                    # print each course from course.__dict__
                    courses = val[term]
                    for course in courses:
                        for i, (c_prop, c_val) in enumerate(course.__dict__.items(), 1):
                            if c_prop == 'grade' and c_val:
                                c_val = float(c_val)
                            if not c_val:
                                c_val = '-'
                            ws.cell(row=r, column=c+i, value=c_val).alignment = Alignment(horizontal='left')
                        r += 1
                        min_rows -= 1
                break
            #print other data
            else:
                ws.cell(row=r, column=c, value=val)
        # pad rows
        while min_rows > 0:
            r+=1
            min_rows-=1
            
        if not export_all:
            r += 1
            
    # Format columns
    for i, (_, w) in enumerate(col_widths.items(), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
#Print college statistics to excel
def export_stats(ws, college, students_dict, term):
    # get stats dict (top_wam, top_sub, avg_wam) for each college
    college_stats, high_perf, under_perf = get_statistics(students_dict, term)

    r = 1
    c = 1
    #print term name
    ws.column_dimensions['A'].width = 22
    ws.cell(row=r, column=c, value=f'{term}: {college_names[college].upper()}').font = Font(bold=True, underline='single')
    r += 1
    
    #if no students that term, exit
    if college_stats["avg_wam"] is None:
        ws.cell(row=r, column=c, value = 'No residents in this term').font = Font(italic=True)
        return
        
    headers = ['Highest Term WAM', 'Best Subject', 'Honorable Mentions (2 or more HDs)']
            
    # print college name
    ws.cell(row=r, column=c, value = 'Average WAM:').font = Font(bold=True)
    ws.cell(row=r, column=c+1, value = f'{college_stats["avg_wam"]:.2f}') 
    r += 2
    ws.cell(row=r, column=c, value = 'High Performers:').font = Font(bold=True)
    r += 1
    for i, (category, tuples) in enumerate(high_perf.items()):
        # print category headers
        r += 1
        ws.cell(row=r, column=c, value=headers[i]).font = Font(underline='single', italic=True)
        r += 1
        # print category data
        for tupl in tuples:
            zid = tupl[0]
            name = students_dict[zid].first_names + ' ' + students_dict[zid].last_name
            
            ws.cell(row=r, column=c, value=name) # print name
            c += 1
            if category == 'top_wam':
                ws.cell(row=r, column=c, value=tupl[1]).alignment = Alignment(horizontal='left') # print wam
            elif category == 'top_sub':
                ws.cell(row=r, column=c, value=f'{tupl[2]}').alignment = Alignment(horizontal='left') # print mark
                ws.cell(row=r, column=c+1, value=f'{tupl[1]}').alignment = Alignment(horizontal='left') # print subject
            elif category == 'full_hd':
                ws.cell(row=r, column=c, value=f'{tupl[1]}/{tupl[2]}').alignment = Alignment(horizontal='left') # print hd/total
            r += 1
            c = 1
            
    r += 1            
    ws.cell(row=r, column=c, value = 'Underperformers:').font = Font(bold=True)
    r += 1            
        
    for zid, courses in under_perf.items():
        student = students_dict[zid]
        name = student.first_names + ' ' + student.last_name
        wam = student.wams[term]
        
        r += 1            
        ws.cell(row=r, column=c, value=name)
        ws.cell(row=r, column=c+1, value = f'WAM: {wam}')
        
        for course in courses:
            for i, (c_prop, c_val) in enumerate(course.__dict__.items(), 2):
                if not c_val:
                    c_val = '-'
                ws.cell(row=r, column=c+i, value=c_val)
                if c_val.upper() in ("FAIL", "ABSENT FAIL", "UNSATISFACTORY FAIL", "ACADEMIC WITHDRAWAL"):
                    ws.cell(row=r, column=c+i).font = Font(color='FF0000')
            r += 1
            
#Iterate through all colleges and print to Excel
def export_to_excel(filename, college_data, term):    
    #create "ALL" data entry
    college_data["ALL"] = {}
    
    # overwrite flag
    # overwrite = 'N'
    
    if os.path.isfile(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        del wb["Sheet"]
    
    for college in college_data:   
        if college != "ALL":
            college_data["ALL"].update(college_data[college])
        
        print (f"Exporting {college}...")
        
        #get worksheet object for college
        if college in wb.sheetnames:
            # if overwrite.lower() != '-y':
            #     overwrite = input(f"Sheet '{college}' already exists, do you want to overwrite? Y/N (or -Y to overwrite all): ")
            #     if overwrite.lower() not in ('y', '-y'):
            #         continue
            del wb[college]
        ws = wb.create_sheet(college)
        ws.sheet_properties.tabColor = college_colours[college] 
        ws.title = college
        
        export_all = True if college == "ALL" else False    
        export_data(ws, college_data[college], term, export_all)
        
        if not export_all:
            # Individual College statistics:
                       
            stats_ws = college + "_stats"
            if stats_ws in wb.sheetnames:
                del wb[stats_ws]
            ws = wb.create_sheet(stats_ws)
            ws.sheet_properties.tabColor = college_colours[college] 
            ws.title = stats_ws
            
            export_stats(ws, college, college_data[college], term)
            
        
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(filename)
    
# Processes a students dict (for a particular college) and returns statistics dict    
# returns tuple of dicts (college_stats, high_perf, under_perf)
def get_statistics(students, term):
    college_stats = {}
    high_perf = {} # dict of lists of 
    under_perf = {}
    
    # college_stats
    total_wam = 0
    wam_count = 0
    
    # high_perf
    # lists of tuples for highest (zid, wam), (zid, code+subject, mark), (zid, hd_count, satis_count)
    top_wam = [('<None>', 0)]
    top_sub = [('<None>', '<None>', 0)]
    # fulfilling 2/2 or 3/3 HDs
    full_hd = []
    also_wam = None #get 2nd best subjects prizes if zid is also highest wam
    
        
    for zid in students:
        if term not in students[zid].terms.keys():
            continue
        t_wam = top_wam[0][1]
        wam = students[zid].wams[term]
        if wam:
            #check if any subjects are higher, and add to top_sub if so        
            wam = float(wam)
            #check is wam higher, and add to top_wam if so        
            if wam > t_wam:
                top_wam.clear()
            if wam >= t_wam:
                top_wam.append((zid, wam))
                
            # for average college wam
            wam_count += 1
            total_wam += wam
        
        #keep track of number of HDs or fails
        hd_count = 0
        fail_count = 0
        for sub in students[zid].terms[term]:
            t_sub = top_sub[0][2] 
            
            if sub.hasGrade():
                grade = int(sub.grade)
                if grade >= t_sub:
                    if zid in [item for tup in top_wam for item in tup]: #if zid also highest wam
                        if not also_wam or grade > also_wam[2]:
                            also_wam = (zid, sub.code + ' ' + sub.name, grade)
                    else:
                        if grade > t_sub:
                            top_sub.clear()
                        top_sub.append((zid, sub.code + ' ' + sub.name, grade))
                        
            if sub.grade_name.upper() == "HIGH DISTINCTION":
                hd_count += 1
            elif sub.grade_name.upper() in ("FAIL", "ABSENT FAIL", "UNSATISFACTORY FAIL", "ACADEMIC WITHDRAWAL"):
                fail_count += 1
                
        sub_count = len(students[zid].terms[term]) #total subjects taken
                
        #add to full_hd if 2 or more HDs
        # if hd_count + sy_count == len(students[zid].terms[term]) \
        if hd_count >= 2:
            full_hd.append((zid, hd_count, sub_count))
        
        #flag if they have failed or is sitting on a Pass   
        if fail_count or (wam and wam < 60):
            under_perf[zid] = students[zid].terms[term]
    
    if top_sub[0][0] == '<None>':
        top_sub.clear()
    if also_wam and also_wam[0] in [item for tup in top_wam for item in tup]: #if also_wam person is in top_wam
        top_sub.insert(0, also_wam)
    
    college_stats = {'avg_wam': total_wam/wam_count} if wam_count > 0 else {'avg_wam': None}
    high_perf = {'top_wam': top_wam, 'top_sub': top_sub, 'full_hd': full_hd}    
        
    return college_stats, high_perf, under_perf



########
# MAIN #
########
def main():
    # Usage
    if len(sys.argv) != 2:
        #default directory is "data/"
        if os.path.isdir("data"):
            directory = "data"
        else:
            print(f"Usage: {sys.argv[0]} <dirname>")
            sys.exit(1)
    else:
        directory = sys.argv[1]
        if not os.path.isdir(directory):
            print(f"Could not find directory '{directory}'")
            sys.exit(1)
            
    # dict of college:data
    college_data = {}
    # dict of all available terms in the data
    all_terms = set()
    
    # last modified times of all files in data dir
    last_mod_load = []
    last_mod = []
    for file in sorted(os.listdir(directory)):
        if file.endswith(".pdf"):
            last_mod.append(os.path.getmtime(f'{directory}/{file}'))
    
    #get any cached data
    cached_dict = f'cache/data.pkl'
    cache_error = False
    
    # CACHING
    # Try to get cached college data
    try:
        with open(f'{cached_dict}', 'rb') as f:
            last_mod_load = pickle.load(f)
            college_data = pickle.load(f)
            all_terms = pickle.load(f)
        
        # check last modified time of data files, if not same, process and repickle
        if last_mod_load != last_mod:
            cache_error = True   
            raise()
        
        print("CACHED DATA WAS FOUND. PDF file processing skipped.\n")
        print("Please delete cache/data.pkl to refresh the cache.\n")
    except Exception:
        cache_error = True
        # Convert files to txt
        pdf_to_txt(directory)
            
        # Loop through all files in curr dir and get data
        for file in sorted(os.listdir(directory)):
            if not file.endswith('.txt'):
                continue
            # College name
            college = file.split()[0]
            college = parse_college_name(college)
            
            # read line by line, generate dict of student objects
            with open(f"{directory}/{file}", 'r') as fp:
                print(f"Processing {college}...")                  
                college_data[college] = parse_lines(fp, college, all_terms)
                
    # precalculate all wams
    for c in college_data:
        for s in college_data[c]:
            college_data[c][s].process_wams()
    
    # User to select term(s)
    terms = pick_term(all_terms)
        
    # process all nominated terms
    while terms:
        term = terms.pop(0)
        
        term_code = convert_term_name(term)
        
        # excel filename
        filename = "College_Academics_" + term_code + ".xlsx"
        
        # Export to excel
        export_to_excel(filename, college_data, term)
        
        print (f"\nDone. File is located at: {os.getcwd()}/{filename}\n")
    
    if not os.path.isfile(cached_dict) or cache_error:
        os.makedirs('cache/', exist_ok=True)
        with open(f'cache/data.pkl', 'wb') as f:
            pickle.dump(last_mod, f)
            pickle.dump(college_data, f)
            pickle.dump(all_terms, f)
            
if __name__ == "__main__":
    main()